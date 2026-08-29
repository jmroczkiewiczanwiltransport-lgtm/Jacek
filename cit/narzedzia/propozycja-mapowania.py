#!/usr/bin/env python3
"""Propozycja mapowania planu kont na znaczniki MF (JPK_KR_PD, węzeł ZOiS).

Prototyp reguł. Docelowo te same reguły trafiają do narzędzia w przeglądarce —
tu liczymy je w Pythonie, bo szybciej sprawdzić trafność na prawdziwym planie kont.

Wejście:  plan kont .xlsx (kolumna 1 = numer konta, kolumna 2 = nazwa)
Wyjście:  .xlsx z propozycją do sprawdzenia przez księgową + statystyka na stdout

Uruchomienie:
  python3 propozycja-mapowania.py PLAN_KONT.xlsx PROPOZYCJA.xlsx [profil] [dziennik.csv]

Podanie dziennika jest zalecane: dorzuca konta, które są w księgach, a nie ma
ich w planie kont, i daje regułom drugą wersję nazwy konta (eksporty
z zagranicznych ERP bywają w innym języku niż plan kont).

profil: erp_zagraniczny_a (domyślny) albo pl_standard — decyduje, jaki zespół
kont odpowiada jakiej grupie bilansowej. Zły profil = złe podpowiedzi, dlatego
narzędzie wypisuje wykryty układ do potwierdzenia.
"""
import difflib
import json
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BAZA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "slowniki")

# Nazwy kont w eksportach z systemów mainframe/ERP przychodzą w rozsypanym
# kodowaniu. Tabela wyprowadzona z kontekstu (patrz cit/PLAN.md).
NAPRAWA_ZNAKOW = {"¡": "Ś", "£": "Ą", "©": "Ż", "¬": "Ł", "Ê": "Ę", "ù": "Ń"}


def napraw_nazwe(s):
    return "".join(NAPRAWA_ZNAKOW.get(c, c) for c in str(s or ""))


def wczytaj_slowniki():
    z = json.load(open(os.path.join(BAZA, "znaczniki-KR_PD.json"), encoding="utf-8"))
    r = json.load(open(os.path.join(BAZA, "reguly-mapowania.json"), encoding="utf-8"))
    opisy = {}
    for typ in ("TMapKontaPOZ", "TMapKontaPD"):
        for x in z["slowniki"][typ]:
            opisy[x["kod"]] = x["opis"]
    sprawdz_reguly(r, z)
    return z, r, opisy


def sprawdz_reguly(reguly, slowniki):
    """Znacznik z reguły, którego nie ma w słowniku MF, dałby plik odrzucony
    przez schemę. Lepiej wysypać się tutaj niż wypuścić taki plik."""
    poz = {x["kod"] for x in slowniki["slowniki"]["TMapKontaPOZ"]}
    pd = {x["kod"] for x in slowniki["slowniki"]["TMapKontaPD"]}
    bledy = []
    for r in reguly["reguly_s121"]:
        if r["znacznik"] not in poz and r["znacznik"] != "?":
            bledy.append("S_12_1: %s (reguła: %s)" % (r["znacznik"], r["nazwa_regex"]))
    for r in reguly["reguly_pd"]:
        if r["znacznik"] not in pd:
            bledy.append("S_12_3: %s (reguła: %s)" % (r["znacznik"], r["nazwa_regex"]))
    if bledy:
        raise SystemExit("Znaczniki nieistniejące w słowniku MF:\n  " + "\n  ".join(bledy))


def wczytaj_plan(sciezka):
    ws = openpyxl.load_workbook(sciezka, data_only=True).active
    konta = []
    for i, w in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if not w or not w[0]:
            continue
        kod = str(w[0]).strip()
        if i == 1 and not re.match(r"^[0-9]", kod):
            continue  # wiersz nagłówka
        konta.append({"wiersz": i, "konto": kod,
                      "nazwa": napraw_nazwe(w[1] if len(w) > 1 else ""), "nazwa2": ""})
    return konta


def dolacz_dziennik(konta, sciezka):
    """Uzupełnia listę kont o te, które są w księgach, a nie ma ich w planie,
    i dokłada nazwę z dziennika jako drugie źródło dla reguł."""
    import csv
    z_dziennika = {}
    for r in csv.DictReader(open(sciezka, encoding="utf-8")):
        z_dziennika.setdefault(r["CPTGE"], r["LICOFD"] or "")
    znane = {k["konto"] for k in konta}
    for k in konta:
        k["nazwa2"] = z_dziennika.get(k["konto"], "")
    for kod in sorted(set(z_dziennika) - znane):
        konta.append({"wiersz": 0, "konto": kod, "nazwa": z_dziennika[kod],
                      "nazwa2": "", "spoza_planu": True})
    return konta


def grupa_konta(kod, profil):
    return profil["zespoly"].get(kod[:1], "")


def dopasuj(reguly, nazwy, grupa):
    """Pierwsza reguła pasująca do którejkolwiek wersji nazwy konta."""
    for r in reguly:
        wymagana = r.get("grupa")
        if wymagana and wymagana != "*" and wymagana != grupa:
            continue
        if any(n and re.search(r["nazwa_regex"], n, re.I) for n in nazwy):
            return r
    return None


def kontekst_wynikowy(grupa):
    if grupa == "koszty":
        return "koszty"
    if grupa == "przychody":
        return "przychody"
    return ""


PRZEDROSTKI_KOREKTA = re.compile(
    r"^(A-|UMORZ\w*\.?|UMORZENIE|ODPIS\w*|ODPIS AKTUAL\.?|INV RESERVE|"
    r"AKUMULOWANA AMORTYZACJA)[\s.\-]*", re.I)
SLOWA_POMIN = re.compile(r"^(FG|PARTS|W|Z|NA|DO|I|TYT|UTRATY)$", re.I)

# Skróty i synonimy z planów kont — bez ich rozwinięcia konto korygujące
# („ODPIS AKTUAL.CZ.KSA") nie trafia na konto wartości („MAGAZYN CZESCI KSA").
SKROTY = {
    "CZ": "CZESCI", "MASZ": "MASZYN", "MASZYNY": "MASZYN", "SR": "SRODKI",
    "SRODKOW": "SRODKI", "TRWALY": "TRWALE", "TRWALYCH": "TRWALE", "TRWA": "TRWALE",
    "TRWAL": "TRWALE", "URZA": "URZADZENIA", "URZ": "URZADZENIA",
    "OB": "OBIEKTY", "INZ": "INZYNIERII", "POZOST": "POZOSTALE", "MAGAZYN": "",
    "AKTUAL": "", "ENEREGET": "ENERGETYCZNE", "ENERGETYCZ": "ENERGETYCZNE",
    "SPECJ": "SPECJALNE", "OGOL": "OGOLNE", "OGÓL": "OGOLNE", "APARATY": "APARATY",
    # należność/zobowiązanie a nazwa konta rozrachunkowego
    "NALEZNOSCI": "ODBIORCAMI", "NALEZNOSC": "ODBIORCAMI",
    "ZOBOWIAZANIA": "DOSTAWCAMI", "KRAJ": "KRAJOWI", "KRAJOWYCH": "KRAJOWI",
    "ZAGRAN": "ZAGRANICA", "ZAGRANICZ": "ZAGRANICA", "UE": "UNIA",
    "LICENCJE": "PATENTY", "PRAWN": "PRAWA", "PRA": "PRAWA", "NIEMAT": "NIEMATERIALNE",
}


def rdzen_nazwy(nazwa):
    """Nazwa konta bez przedrostka korekty i szumu, ze rozwiniętymi skrótami."""
    n = PRZEDROSTKI_KOREKTA.sub("", nazwa.upper())
    n = re.sub(r"[^A-ZĄĆĘŁŃÓŚŹŻ0-9]+", " ", n)
    slowa = []
    for w in n.split():
        if SLOWA_POMIN.match(w):
            continue
        w = SKROTY.get(w, w)
        if w:
            slowa.append(w)
    return slowa


def podobienstwo(a, b):
    """Zbieżność dwóch rdzeni: większa z miary sekwencyjnej i pokrycia słów.

    Sama miara sekwencyjna gubi przestawione słowa („LICENCJE,PATENTY" vs
    „PATENTY,MARKI"), samo pokrycie słów myli konta różniące się jednym członem.
    """
    if not a or not b:
        return 0.0
    seq = difflib.SequenceMatcher(None, " ".join(a), " ".join(b)).ratio()
    za, zb = set(a), set(b)
    pokrycie = len(za & zb) / float(len(za | zb))
    ocena = max(seq, pokrycie)
    # zaliczka to nie ten sam składnik majątku co rzecz, której dotyczy
    if ("ZALICZKI" in za) != ("ZALICZKI" in zb):
        ocena -= 0.15
    # Krótkie kody (KSA, KHU, JD) to segmenty albo marki. Jeśli kod z jednej
    # nazwy nie występuje w drugiej, para jest wątpliwa mimo podobnego brzmienia
    # („MASZ.KHU" vs „MASZYN KAU" różnią się jedną literą, a to inny magazyn).
    kod = re.compile(r"^[A-Z]{2,4}$")
    kody_a = {w for w in za if kod.match(w)}
    kody_b = {w for w in zb if kod.match(w)}
    if kody_a and kody_b and not (kody_a & kody_b):
        ocena = min(ocena, 0.7)
    return ocena


def sparuj_korekty(wynik, opisy):
    """Konto umorzeniowe / odpisowe dziedziczy znacznik konta wartości brutto,
    z końcówką _U (umorzenie) albo _A (odpis aktualizujący).

    Parujemy po rdzeniu nazwy, wspierając się zbieżnością numeru konta —
    w planach kont konto korygujące zwykle powtarza końcówkę konta wartości.
    """
    zrodla = [w for w in wynik
              if w["s121"] not in ("?",) and not PRZEDROSTKI_KOREKTA.match(w["nazwa"].upper())]
    for w in wynik:
        if w["s121"] != "?":
            continue
        gora = w["dlaczego"]
        koncowka = "_U" if "umorzenio" in gora else ("_A" if "odpis" in gora else None)
        if not koncowka:
            continue
        rdzen = rdzen_nazwy(w["nazwa"])
        if not rdzen:
            continue
        najlepsze, ocena_najlepsza = None, 0.0
        for z in zrodla:
            ocena = podobienstwo(rdzen, rdzen_nazwy(z["nazwa"]))
            if ocena <= 0:
                continue
            # zbieżność numeru konta: wspólna końcówka podnosi ocenę
            wspolna = 0
            for a, b in zip(reversed(w["konto"]), reversed(z["konto"])):
                if a != b:
                    break
                wspolna += 1
            ocena += 0.05 * wspolna
            if ocena > ocena_najlepsza:
                najlepsze, ocena_najlepsza = z, ocena
        if not najlepsze or ocena_najlepsza < 0.5:
            continue
        kandydat = bazowy_znacznik(najlepsze["s121"]) + koncowka
        if kandydat not in opisy:
            w["dlaczego"] += (" | para: %s %s — ale MF nie przewiduje wariantu %s dla tej pozycji"
                              % (najlepsze["konto"], najlepsze["nazwa"], koncowka))
            continue
        w["s121"] = kandydat
        w["opis_s121"] = opisy[kandydat]
        # Przy słabym dopasowaniu nazw para bywa z sąsiedniego segmentu — pozycja
        # bilansowa zwykle wychodzi ta sama, ale księgowa musi to zobaczyć.
        pewne = ocena_najlepsza >= 0.75
        w["pewnosc"] = "srednia" if pewne else "niska"
        w["dlaczego"] = ("znacznik odziedziczony z konta wartości brutto %s (%s) z końcówką %s — %s"
                         % (najlepsze["konto"], najlepsze["nazwa"], koncowka,
                            "potwierdzić, że to właściwa para" if pewne else
                            "UWAGA: nazwy nie zgadzają się dokładnie, para dobrana najbliższą "
                            "dostępną — sprawdzić obowiązkowo"))


def oznacz_powiazane(wynik, reguly, opisy):
    """Przełącza znacznik na wariant dla jednostek powiązanych.

    Znaczniki MF rozdzielają część pozycji (należności, zobowiązania, sprzedaż,
    wartość sprzedanych towarów, odsetki) na jednostki powiązane i pozostałe.
    Rozpoznajemy je po tokenach z planu kont, które wskazuje klient — nazwa
    spółki z grupy, oznaczenie cash poolingu, skrót segmentu.
    """
    kfg = reguly.get("powiazane")
    if not kfg:
        return
    wzor = re.compile("|".join(
        [re.escape(t) for t in kfg.get("tokeny", [])]
        + [r"\b%s\b" % re.escape(t) for t in kfg.get("tokeny_slowo", [])]), re.I)
    zamiana = kfg.get("zamiana", {})
    bez_podzialu = set(kfg.get("grupy_bez_podzialu", []))
    for w in wynik:
        if w["grupa"] in bez_podzialu or not wzor.search(w["nazwa"]):
            continue
        stary = w["s121"]
        nowy = zamiana.get(stary)
        if not nowy and stary.endswith("_POZ"):
            nowy = stary[:-4] + "_POW"
        if not nowy or nowy not in opisy:
            # pozycja bez wariantu powiązanego (np. koszty rodzajowe, zapasy)
            continue
        w["s121"] = nowy
        w["opis_s121"] = opisy[nowy]
        w["dlaczego"] = ("jednostka powiązana (rozpoznana po nazwie konta) — wariant „%s” "
                         "zamiast „%s”; %s" % (nowy, stary, w["dlaczego"]))


def bazowy_znacznik(kod):
    return re.sub(r"_(W|A|U|O)$", "", kod)


def main():
    wejscie, wyjscie = sys.argv[1], sys.argv[2]
    nazwa_profilu = sys.argv[3] if len(sys.argv) > 3 else "erp_zagraniczny_a"
    _, reguly, opisy = wczytaj_slowniki()
    profil = reguly["profile_planu_kont"][nazwa_profilu]
    konta = wczytaj_plan(wejscie)
    if len(sys.argv) > 4:
        konta = dolacz_dziennik(konta, sys.argv[4])

    wynik = []
    for k in konta:
        grupa = grupa_konta(k["konto"], profil)
        nazwy = [k["nazwa"], k.get("nazwa2", "")]
        tr = dopasuj(reguly["reguly_s121"], nazwy, grupa)
        znacznik = tr["znacznik"] if tr else "?"
        pewnosc = tr["pewnosc"] if tr else "brak"
        dlaczego = tr["dlaczego"] if tr else "żadna reguła nie pasuje — konto do ręcznego przypisania"

        pd, pd_dlaczego = "", ""
        kon = kontekst_wynikowy(grupa)
        for r in reguly["reguly_pd"]:
            if r.get("kontekst") and r["kontekst"] != kon:
                continue
            if any(n and re.search(r["nazwa_regex"], n, re.I) for n in nazwy):
                pd, pd_dlaczego = r["znacznik"], r["dlaczego"]
                break

        wynik.append({
            "konto": k["konto"], "nazwa": k["nazwa"], "grupa": grupa,
            "s121": znacznik, "opis_s121": opisy.get(znacznik, ""),
            "pewnosc": pewnosc, "dlaczego": dlaczego,
            "s123": pd, "opis_s123": opisy.get(pd, ""), "dlaczego_pd": pd_dlaczego,
        })

    sparuj_korekty(wynik, opisy)
    oznacz_powiazane(wynik, reguly, opisy)
    zapisz(wyjscie, wynik, nazwa_profilu, profil)
    statystyka(wynik, profil)


def zapisz(sciezka, wynik, nazwa_profilu, profil):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Propozycja mapowania"
    naglowki = ["Konto", "Nazwa konta", "Grupa (z profilu)", "S_12_1 propozycja",
                "Co to znaczy (MF)", "Pewność", "Dlaczego tak",
                "S_12_3 (PD)", "Co to znaczy (MF)", "Dlaczego tak",
                "POPRAWKA księgowej (S_12_1)", "POPRAWKA księgowej (S_12_3)", "Uwagi księgowej"]
    ws.append(naglowki)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F6E64")
        c.alignment = Alignment(vertical="center", wrap_text=True)

    kolory = {"wysoka": None, "srednia": "FFF6DF", "niska": "FFE6D5", "brak": "FFD5D5"}
    for w in wynik:
        ws.append([w["konto"], w["nazwa"], w["grupa"], w["s121"], w["opis_s121"],
                   w["pewnosc"], w["dlaczego"], w["s123"], w["opis_s123"], w["dlaczego_pd"],
                   "", "", ""])
        kol = kolory.get(w["pewnosc"])
        if kol:
            for c in ws[ws.max_row]:
                c.fill = PatternFill("solid", fgColor=kol)

    for kol, szer in zip("ABCDEFGHIJKLM", [10, 32, 16, 18, 52, 10, 60, 12, 46, 46, 26, 26, 30]):
        ws.column_dimensions[kol].width = szer
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Jak czytać")
    for wiersz in [
        ["Propozycja mapowania kont na znaczniki MF — do sprawdzenia przez księgową"],
        [],
        ["Profil planu kont", nazwa_profilu],
        ["", profil["opis"]],
        [],
        ["S_12_1", "znacznik pozycji bilansu / rachunku wyników. WYMAGANY dla każdego konta (węzeł ZOiS7 — jednostki pozostałe)."],
        ["S_12_2", "drugi znacznik, opcjonalny — gdy jedno konto zasila dwie pozycje."],
        ["S_12_3", "znacznik podatkowy (PD) — opcjonalny, dla kont z różnicami między wynikiem księgowym a podatkowym."],
        [],
        ["Pewność: wysoka", "nazwa konta jednoznacznie wskazuje pozycję"],
        ["Pewność: średnia", "wskazuje, ale są warianty (np. krótko- / długoterminowe, powiązane / pozostałe)"],
        ["Pewność: niska", "podpowiedź na podstawie zespołu kont — sprawdzić obowiązkowo"],
        ["Pewność: brak", "żadna reguła nie pasuje — przypisanie w całości po stronie księgowej"],
        [],
        ["Znacznik „?”", "reguła rozpoznała rodzaj konta, ale wskazanie pozycji wymaga decyzji (np. konto umorzeniowe wymaga pary z kontem wartości brutto)."],
        [],
        ["Co robimy z poprawkami", "kolumny POPRAWKA wypełnia księgowa; wracają do nas i stają się mapowaniem tej firmy — w kolejnych latach wczytuje się gotowe."],
    ]:
        info.append(wiersz)
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 100
    info["A1"].font = Font(bold=True, size=13)
    for r in info.iter_rows(min_col=2, max_col=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(sciezka)


def statystyka(wynik, profil):
    n = len(wynik)
    print("Kont w planie:", n)
    print("Profil:", profil["opis"])
    print("\nPewność propozycji S_12_1:")
    for p in ("wysoka", "srednia", "niska", "brak"):
        ile = sum(1 for w in wynik if w["pewnosc"] == p)
        print("  %-8s %4d  (%.0f%%)" % (p, ile, 100.0 * ile / n))
    znak = sum(1 for w in wynik if w["s121"] not in ("?",))
    print("\nKont z konkretnym znacznikiem: %d (%.0f%%)" % (znak, 100.0 * znak / n))
    print("Kont wymagających decyzji („?” albo brak reguły): %d" % (n - znak))
    print("Kont ze znacznikiem podatkowym PD: %d" % sum(1 for w in wynik if w["s123"]))
    pow_ = [w for w in wynik if "jednostka powiązana" in w["dlaczego"]]
    print("Kont przełączonych na wariant jednostek powiązanych: %d" % len(pow_))
    for w in pow_:
        print("   %-8s %-32s -> %s" % (w["konto"], w["nazwa"][:32], w["s121"]))
    spoza = [w for w in wynik if "nie ma w planie kont" in w["dlaczego"]]
    if spoza:
        print("\nKonta z dziennika spoza planu kont (%d) — do dopisania w planie:" % len(spoza))
        for w in spoza:
            print("   %-8s %-34s -> %s" % (w["konto"], w["nazwa"][:34], w["s121"]))
    braki = [w for w in wynik if w["pewnosc"] == "brak"]
    if braki:
        print("\nKonta bez żadnej reguły (do uzupełnienia reguł albo ręcznie):")
        for w in braki[:40]:
            print("  ", w["konto"], w["nazwa"])
        if len(braki) > 40:
            print("   ... i jeszcze %d" % (len(braki) - 40))


if __name__ == "__main__":
    main()
