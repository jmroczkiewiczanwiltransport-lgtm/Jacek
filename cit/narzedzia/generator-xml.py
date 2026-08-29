#!/usr/bin/env python3
"""Składa plik JPK_KR_PD z dziennika zapisów, planu kont i mapowania znaczników.

Stan: składa węzły Naglowek, Podmiot1, Kontrahent, ZOiS i RPD. Węzeł Dziennik
powstaje tylko wtedy, gdy podano eksport dziennika z identyfikatorem zapisu —
zwykły wydruk księgi głównej (grand livre) nie wystarcza, bo nie ma w nim
numeru dowodu, rodzaju dowodu ani trzech dat wymaganych przez strukturę.
Bez węzła Dziennik plik nie przejdzie walidacji schemą i narzędzie o tym mówi.

Uruchomienie:
  python3 generator-xml.py dziennik.csv plan_kont.xlsx podmiot.json wyjscie.xml
"""
import collections
import csv
import json
import os
import re
import sys
from decimal import Decimal
from xml.sax.saxutils import escape

BAZA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "slowniki")
NS = "http://jpk.mf.gov.pl/wzor/2024/09/04/09041/"
ETD = "http://crd.gov.pl/xml/schematy/dziedzinowe/mf/2022/01/05/eD/DefinicjeTypy/"
ZERO = Decimal("0")
NAPRAWA_ZNAKOW = {"¡": "Ś", "£": "Ą", "©": "Ż", "¬": "Ł", "Ê": "Ę", "ù": "Ń"}


def napraw(s):
    return "".join(NAPRAWA_ZNAKOW.get(c, c) for c in str(s or ""))


def kw(x):
    """Pole kwotowe: 2 miejsca po przecinku, kropka dziesiętna."""
    return str(Decimal(x).quantize(Decimal("0.01")))


def el(nazwa, wartosc, wciecie=6):
    return "%s<%s>%s</%s>" % (" " * wciecie, nazwa, escape(str(wartosc)), nazwa)


def wczytaj_dziennik(sciezka):
    rows = list(csv.DictReader(open(sciezka, encoding="utf-8")))
    for r in rows:
        r["kwota"] = Decimal(r["MTLIG"]) if r["MTLIG"] else ZERO
    return rows


def wczytaj_plan(sciezka):
    import openpyxl
    ws = openpyxl.load_workbook(sciezka, data_only=True).active
    plan = {}
    for w in ws.iter_rows(values_only=True):
        if not w or not w[0]:
            continue
        kod = str(w[0]).strip()
        if kod[:1].isdigit():
            plan[kod] = napraw(w[1] if len(w) > 1 else "")
    return plan


def wczytaj_mapowanie(sciezka):
    """Mapowanie kont na znaczniki — z arkusza sprawdzonego przez księgową.

    Kolumny POPRAWKA mają pierwszeństwo przed propozycją: to, co księgowa
    wpisała ręcznie, zawsze wygrywa z tym, co zaproponowały reguły.
    """
    import openpyxl
    ws = openpyxl.load_workbook(sciezka, data_only=True).active
    mapa = {}
    for w in ws.iter_rows(min_row=2, values_only=True):
        if not w or not w[0]:
            continue
        kod = str(w[0]).strip()
        s121 = (w[10] or w[3] or "").strip()
        s123 = (w[11] or w[7] or "").strip()
        mapa[kod] = (s121 if s121 != "?" else "", s123)
    return mapa


def salda(dziennik):
    bo_wn = collections.defaultdict(lambda: ZERO)
    bo_ma = collections.defaultdict(lambda: ZERO)
    ob_wn = collections.defaultdict(lambda: ZERO)
    ob_ma = collections.defaultdict(lambda: ZERO)
    for r in dziennik:
        k, v = r["CPTGE"], r["kwota"]
        if r["ORIGIN"] == "OUV":
            (bo_wn if v >= 0 else bo_ma)[k] += abs(v)
        else:
            (ob_wn if v >= 0 else ob_ma)[k] += abs(v)
    return bo_wn, bo_ma, ob_wn, ob_ma


def konto_nadrzedne(kod, plan):
    """Identyfikator konta nadrzędnego (S_3) — pole wymagane, nie może być puste.

    W planach kont bez zbudowanej hierarchii (a taki jest plan pierwszego
    klienta: płaskie konta sześciocyfrowe) przyjmujemy konto syntetyczne
    z tego samego zespołu, a gdy konto samo nim jest — jego trzycyfrowy
    symbol. Do potwierdzenia przez księgową: to jedyne pole ZOiS, którego
    nie da się wyprowadzić z danych bez decyzji.
    """
    syntetyczne = kod[:3] + "000"
    if syntetyczne != kod and syntetyczne in plan:
        return syntetyczne
    return kod[:3]


def wezel_zois(dziennik, plan, mapa, braki):
    bo_wn, bo_ma, ob_wn, ob_ma = salda(dziennik)
    konta = sorted(set(list(bo_wn) + list(bo_ma) + list(ob_wn) + list(ob_ma)))
    out = []
    for k in konta:
        nazwa = plan.get(k)
        if not nazwa:
            nazwa = next((r["LICOFD"] for r in dziennik if r["CPTGE"] == k and r["LICOFD"]), k)
            braki.append("konto %s nie występuje w planie kont — nazwa wzięta z dziennika" % k)
        s121, s123 = mapa.get(k, ("", ""))
        if not s121:
            braki.append("konto %s (%s) nie ma znacznika S_12_1 — pole jest wymagane" % (k, nazwa))
        bw, bm = bo_wn[k], bo_ma[k]
        ow, om = ob_wn[k], ob_ma[k]
        saldo = (bw + ow) - (bm + om)
        out.append("    <ZOiS7>")
        out.append(el("S_1", k))
        out.append(el("S_2", nazwa[:256]))
        out.append(el("S_3", konto_nadrzedne(k, plan)))
        for nazwa_pola, wartosc in (("S_4", bw), ("S_5", bm), ("S_6", ow), ("S_7", om),
                                    ("S_8", bw + ow), ("S_9", bm + om),
                                    ("S_10", saldo if saldo > 0 else ZERO),
                                    ("S_11", -saldo if saldo < 0 else ZERO)):
            out.append(el(nazwa_pola, kw(wartosc)))
        if s121:
            out.append(el("S_12_1", s121))
        if s123:
            out.append(el("S_12_3", s123))
        out.append("    </ZOiS7>")
    return out, konta


def wezel_rpd(dziennik, mapa):
    """K_1–K_8 z sald kont niosących znacznik podatkowy.

    Do pól JPK trafiają wartości bezwzględne; znak decyduje tylko o tym,
    do której pozycji kwota należy (koszt odwrócony z lat ubiegłych trafia
    do K_6, a nie ujemnie do K_4).
    """
    pd_do_pola = {"PD1": "K_1", "PD1_1": "K_1", "PD1_2": "K_1", "PD1_3": "K_1",
                  "PD2": "K_2", "PD4": "K_4", "PD4_1": "K_4", "PD4_2": "K_4",
                  "PD4_3": "K_4", "PD5": "K_5"}
    odwrocone = {"K_4": "K_6", "K_5": "K_6", "K_1": "K_3", "K_2": "K_3"}
    saldo = collections.defaultdict(lambda: ZERO)
    for r in dziennik:
        saldo[r["CPTGE"]] += r["kwota"]
    pola = collections.defaultdict(lambda: ZERO)
    for kod, (_, s123) in mapa.items():
        pole = pd_do_pola.get(s123)
        if not pole:
            continue
        v = saldo.get(kod, ZERO)
        if v == 0:
            continue
        koszt = pole in ("K_4", "K_5")
        naturalny = (v > 0) if koszt else (v < 0)
        pola[pole if naturalny else odwrocone[pole]] += abs(v)
    return ["    " + el("K_%d" % i, kw(pola["K_%d" % i]), 0).strip() for i in range(1, 9)]


def main():
    dziennik = wczytaj_dziennik(sys.argv[1])
    plan = wczytaj_plan(sys.argv[2])
    podmiot = json.load(open(sys.argv[3], encoding="utf-8"))
    wyjscie = sys.argv[4]
    mapa = wczytaj_mapowanie(podmiot["mapowanie"]) if podmiot.get("mapowanie") else {}

    braki = []
    zois, konta = wezel_zois(dziennik, plan, mapa, braki)
    rpd = wezel_rpd(dziennik, mapa)

    kontrahenci = sorted({r["CDCLI"] for r in dziennik if r["CDCLI"] and r["CDCLI"] != "0"})
    adres = podmiot["adres"]
    x = ['<?xml version="1.0" encoding="UTF-8"?>',
         '<JPK xmlns="%s" xmlns:etd="%s">' % (NS, ETD),
         "  <Naglowek>",
         '    <KodFormularza kodSystemowy="JPK_KR_PD (1)" wersjaSchemy="1-1">JPK_KR_PD</KodFormularza>',
         el("WariantFormularza", 1, 4),
         el("CelZlozenia", podmiot.get("cel_zlozenia", 1), 4),
         el("DataWytworzeniaJPK", podmiot["data_wytworzenia"], 4),
         el("DataOd", podmiot["data_od"], 4), el("DataDo", podmiot["data_do"], 4),
         el("RokDataOd", podmiot["rok_od"], 4), el("RokDataDo", podmiot["rok_do"], 4),
         el("DomyslnyKodWaluty", podmiot.get("waluta", "PLN"), 4),
         el("KodUrzedu", podmiot["kod_urzedu"], 4),
         "  </Naglowek>", "  <Podmiot1>", "    <IdentyfikatorPodmiotu>",
         el("etd:NIP", podmiot["nip"], 6), el("etd:PelnaNazwa", podmiot["nazwa"], 6),
         "    </IdentyfikatorPodmiotu>", "    <Adres>", "      <AdresPol>",
         el("etd:KodKraju", adres.get("kraj", "PL"), 8),
         el("etd:Wojewodztwo", adres["wojewodztwo"], 8),
         el("etd:Powiat", adres["powiat"], 8), el("etd:Gmina", adres["gmina"], 8),
         el("etd:NrDomu", adres["nr_domu"], 8),
         el("etd:Miejscowosc", adres["miejscowosc"], 8),
         el("etd:KodPocztowy", adres["kod_pocztowy"], 8),
         "      </AdresPol>", "    </Adres>", "  </Podmiot1>"]
    for k in kontrahenci:
        x += ["  <Kontrahent>", el("T_1", k, 4), "  </Kontrahent>"]
    x += ["  <ZOiS>"] + zois + ["  </ZOiS>"]
    if podmiot.get("dziennik_xml"):
        x.append(open(podmiot["dziennik_xml"], encoding="utf-8").read().rstrip())
    else:
        braki.append("brak węzła Dziennik — potrzebny eksport dziennika zapisów "
                     "(numer zapisu, numer i rodzaj dowodu, trzy daty, osoba odpowiedzialna)")
    x += ["  <RPD>"] + rpd + ["  </RPD>", "</JPK>"]
    open(wyjscie, "w", encoding="utf-8").write("\n".join(x) + "\n")

    print("Zapisano %s" % wyjscie)
    print("Kont w ZOiS: %d | kontrahentów: %d" % (len(konta), len(kontrahenci)))
    print("RPD: %s" % " ".join(re.sub(r"\s+", "", p) for p in rpd))
    if braki:
        print("\nCZEGO BRAKUJE, ŻEBY PLIK BYŁ KOMPLETNY (%d):" % len(braki))
        for b in braki[:15]:
            print("  - %s" % b)
        if len(braki) > 15:
            print("  ... i jeszcze %d" % (len(braki) - 15))


if __name__ == "__main__":
    main()
